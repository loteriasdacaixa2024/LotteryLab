import sqlite3
import os
import time
import logging
import multiprocessing
import math
from itertools import combinations, product
from typing import List, Tuple, Dict, Optional, Any, Union
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    import csv
    OPENPYXL_AVAILABLE = False

# Importações Rich opcionais (mantidas apenas para possível uso futuro nos relatórios se necessário)
# mas removidas da execução principal por preferência do usuário

# Configuração de Log para console
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def number_list_to_bitmask(numbers: List[int]) -> int:
    mask = 0
    for n in numbers:
        if n > 0:
            mask |= (1 << (n - 1))
    return mask

def bitmask_to_numbers(mask: int, universe_size: int) -> List[int]:
    return [i + 1 for i in range(universe_size) if (mask >> i) & 1]

def worker_task_positional(args):
    """
    Worker para loterias posicionais (Super Sete).
    Compara index a index.
    """
    try:
        prefix, universe_size, numbers_to_pick, contest_data, min_hits, min_occ, sec_hits, sec_occ = args
        results = []
        
        remaining_to_pick = numbers_to_pick - len(prefix)
        # Para posicional, o pool é sempre 0-9 (universe_size)
        remaining_pool = range(universe_size)
        
        limit_hits = numbers_to_pick + 1
        
        for rest in product(remaining_pool, repeat=remaining_to_pick):
            comb = list(prefix) + list(rest)
            
            max_hits = 0
            hits_count = [0] * limit_hits
            
            for target in contest_data:
                # Compara cada posição
                hits = sum(1 for i in range(numbers_to_pick) if comb[i] == target[i])
                if hits < limit_hits:
                    hits_count[hits] += 1
                if hits > max_hits:
                    max_hits = hits
            
            meets = False
            if min_hits < limit_hits and hits_count[min_hits] >= min_occ:
                meets = True
            elif sec_hits and sec_hits < limit_hits and hits_count[sec_hits] >= sec_occ:
                meets = True
                
            if meets:
                results.append((comb, max_hits, hits_count[min_hits], (hits_count[sec_hits] if sec_hits and sec_hits > 0 else 0)))
        return results
    except Exception as e:
        import traceback
        return f"Error: {str(e)}\n{traceback.format_exc()}"

def worker_task(args):
    """
    Worker otimizado para Windows (Top-level function para Pickling).
    Processa combinações com base em prefixos fixos para evitar overhead de islice.
    """
    try:
        prefix, universe_size, numbers_to_pick, contest_masks, min_hits, min_occ, sec_hits, sec_occ = args
        results = []
        
        # O prefixo é uma tupla, ex: (1, 2)
        prefix_mask = 0
        for n in prefix:
            prefix_mask |= (1 << (n - 1))
            
        remaining_to_pick = numbers_to_pick - len(prefix)
        # O pool de números restantes deve ser maior que o último número do prefixo
        remaining_pool = range(prefix[-1] + 1, universe_size + 1)
        
        limit_hits = max(min_hits, (sec_hits if sec_hits else 0)) + 1
        
        # Se não houver números suficientes para completar a combinação, retorna vazio
        if remaining_to_pick < 0: return []
        
        # Gera combinações parciais
        for rest in combinations(remaining_pool, remaining_to_pick):
            comb_mask = prefix_mask
            for n in rest:
                comb_mask |= (1 << (n - 1))
                
            max_hits = 0
            hits_count = [0] * limit_hits
            
            for cont_mask in contest_masks:
                hits = (comb_mask & cont_mask).bit_count()
                if hits < limit_hits:
                    hits_count[hits] += 1
                if hits > max_hits:
                    max_hits = hits
            
            # Critério de salvamento
            meets = False
            if hits_count[min_hits] >= min_occ:
                meets = True
            elif sec_hits and sec_hits > 0 and hits_count[sec_hits] >= sec_occ:
                meets = True
                
            if meets:
                results.append((comb_mask, max_hits, hits_count[min_hits], (hits_count[sec_hits] if sec_hits and sec_hits > 0 else 0)))
        return results
    except Exception as e:
        return f"Error: {str(e)}"

def worker_task_candidate(args):
    """
    Worker para o Modo Turbo (Candidatos).
    """
    try:
        candidate_masks, universe_size, numbers_to_pick, contest_masks, min_hits, min_occ, sec_hits, sec_occ = args
        results = []
        limit_hits = max(min_hits, (sec_hits if sec_hits else 0)) + 1
        
        for comb_mask in candidate_masks:
            max_hits = 0
            hits_count = [0] * limit_hits
            
            for cont_mask in contest_masks:
                hits = (comb_mask & cont_mask).bit_count()
                if hits < limit_hits:
                    hits_count[hits] += 1
                if hits > max_hits:
                    max_hits = hits
            
            meets = False
            if hits_count[min_hits] >= min_occ:
                meets = True
            elif sec_hits and sec_hits > 0 and hits_count[sec_hits] >= sec_occ:
                meets = True
                
            if meets:
                results.append((comb_mask, max_hits, hits_count[min_hits], (hits_count[sec_hits] if sec_hits and sec_hits > 0 else 0)))
        return results
    except Exception as e:
        return f"Error: {str(e)}"

class BacktestEngine:
    def __init__(self, config):
        self.config = config
        self.db_path = os.path.join('bancos', config.db_file)
        self.contest_masks = []
        
    def load_data(self):
        logger.info(f"Conectando ao banco histórico: {self.db_path}")
        if not os.path.exists(self.db_path):
            raise FileNotFoundError(f"Banco de dados {self.db_path} não encontrado.")
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        first_elem = self.config.ball_columns[0]
        column_sets = self.config.ball_columns if isinstance(first_elem, list) else [self.config.ball_columns]
            
        all_cols = [c for s in column_sets for c in s]
        unique_cols = list(dict.fromkeys(all_cols))
        query = f"SELECT {', '.join(unique_cols)} FROM {self.config.table_name}"
        
        try:
            cursor.execute(query)
            rows = cursor.fetchall()
            col_to_idx = {col: i for i, col in enumerate(unique_cols)}
            for row in rows:
                for s in column_sets:
                    nums = []
                    for col in s:
                        val = row[col_to_idx[col]]
                        if val is None:
                            continue
                        if isinstance(val, str) and ',' in val:
                            try:
                                nums.extend([int(x.strip()) for x in val.split(',')])
                            except ValueError:
                                pass
                        else:
                            nums.append(val)
                    
                    if nums:
                        if self.config.is_positional:
                            self.contest_masks.append(nums)
                        else:
                            self.contest_masks.append(number_list_to_bitmask(nums))
            logger.info(f"Dados carregados: {len(self.contest_masks)} sorteios históricos processados.")
        finally:
            conn.close()

    def run(self, num_processes: Optional[int] = None):
        if not self.contest_masks:
            self.load_data()
            
        if self.config.is_positional:
            total_comb = self.config.universe_size ** self.config.numbers_to_pick
        else:
            total_comb = math.comb(self.config.universe_size, self.config.numbers_to_pick)
        start_time_global = time.time()
        logger.info(f"Iniciando Motor de Backtesting: {self.config.name}")
        logger.info(f"Total teórico de combinações: {total_comb:,}")
        
        num_processes = num_processes or multiprocessing.cpu_count()
        logger.info(f"Utilizando {num_processes} threads de processamento.")

        if total_comb > 20_000_000:
            logger.info("⚠️ Espaço amostral gigante detectado! Ativando Modo Candidato Inteligente (Turbo)...")
            final_data = self.run_smart_candidate_mode(num_processes)
        elif self.config.is_positional:
            final_data = self.run_positional_mode(num_processes)
        else:
            final_data = self.run_brute_force_mode(num_processes)
            
        elapsed = time.time() - start_time_global
        logger.info(f"Processo completo em {elapsed:.2f} segundos.")
        return final_data

    def run_brute_force_mode(self, num_processes):
        start_time = time.time()
        logger.info("Fatiando espaço amostral para feedback em tempo real...")
        task_args_list = []
        
        # Divisão por prefixos de 3 números para feedback
        for n1 in range(1, self.config.universe_size - self.config.numbers_to_pick + 2):
            for n2 in range(n1 + 1, self.config.universe_size - self.config.numbers_to_pick + 3):
                for n3 in range(n2 + 1, self.config.universe_size - self.config.numbers_to_pick + 4):
                    prefix = (n1, n2, n3)
                    task_args = (
                        prefix, 
                        self.config.universe_size, 
                        self.config.numbers_to_pick,
                        self.contest_masks,
                        self.config.min_hits_to_save,
                        self.config.min_occurences,
                        self.config.secondary_hits,
                        self.config.secondary_min_occurences
                    )
                    task_args_list.append(task_args)
            
        total_tasks = len(task_args_list)
        logger.info(f"Processamento distribuído em {total_tasks} micro-blocos.")
        
        final_data = []
        last_log = start_time
        
        pool = multiprocessing.Pool(processes=num_processes)
        try:
            for i, batch_res in enumerate(pool.imap_unordered(worker_task, task_args_list)):
                completed = i + 1
                if batch_res and not isinstance(batch_res, str):
                    final_data.extend(batch_res)
                
                now = time.time()
                if now - last_log >= 1 or completed == total_tasks:
                    percent = (completed / total_tasks) * 100
                    logger.info(f"Progresso: {percent:.2f}% | Blocos: {completed}/{total_tasks}")
                    last_log = now
        finally:
            pool.close()
            pool.join()
            
        elapsed = time.time() - start_time
        logger.info(f"Backtest finalizado em {elapsed:.2f} segundos.")
        logger.info(f"Ranking: {len(final_data)} combinações encontradas.")
        if final_data: self.save_results(final_data)
        return final_data

    def run_positional_mode(self, num_processes):
        start_time = time.time()
        logger.info("Iniciando processamento posicional (Repetições permitidas)...")
        task_args_list = []
        
        # Divisão por prefixos de 2 números para 10^7 combinações
        for prefix in product(range(self.config.universe_size), repeat=2):
            task_args = (
                prefix, 
                self.config.universe_size, 
                self.config.numbers_to_pick,
                self.contest_masks,
                self.config.min_hits_to_save,
                self.config.min_occurences,
                self.config.secondary_hits,
                self.config.secondary_min_occurences
            )
            task_args_list.append(task_args)
            
        total_tasks = len(task_args_list)
        logger.info(f"Processamento distribuído em {total_tasks} blocos posicionais.")
        
        final_data = []
        last_log = start_time
        
        pool = multiprocessing.Pool(processes=num_processes)
        try:
            for i, batch_res in enumerate(pool.imap_unordered(worker_task_positional, task_args_list)):
                completed = i + 1
                if batch_res and isinstance(batch_res, list):
                    final_data.extend(batch_res)
                elif isinstance(batch_res, str) and batch_res.startswith("Error"):
                    logger.error(batch_res)
                
                now = time.time()
                if now - last_log >= 1 or completed == total_tasks:
                    percent = (completed / total_tasks) * 100
                    logger.info(f"Progresso: {percent:.2f}% | {completed}/{total_tasks}")
                    last_log = now
        finally:
            pool.close()
            pool.join()
            
        logger.info(f"Ranking: {len(final_data)} combinações encontradas.")
        if final_data: self.save_results(final_data)
        return final_data

    def run_smart_candidate_mode(self, num_processes):
        """
        Modo Turbo: Gera candidatos baseados nos resultados históricos.
        Qualquer combinação que atinja o critério de elite DEVE ser um vizinho de 0 ou 1
        distância de algum sorteio real.
        """
        start_time = time.time()
        logger.info("Identificando candidatos potenciais baseados no histórico...")
        
        candidates = set()
        # 1. Todo resultado histórico é um candidato (0-distância, busca hits_main)
        for mask in self.contest_masks:
            candidates.add(mask)
            
        # 2. Toda combinação que variou 1 número de algum resultado (1-distância, busca hits_sec)
        # Se min_hits_to_save=7 e secondary_hits=6, um 6-hits é um 1-vizinho de um 7-hits.
        logger.info(f"Gerando vizinhança para {len(self.contest_masks)} sorteios...")
        
        # Para cada sorteio, pegamos subconjuntos de (pick - 1) e completamos com todos os outros números
        for mask in self.contest_masks:
            nums = bitmask_to_numbers(mask, self.config.universe_size)
            # Para Timemania (80,7), nums tem 7 números.
            # Subconjuntos de 6:
            for sub in combinations(nums, self.config.numbers_to_pick - 1):
                base_mask = number_list_to_bitmask(list(sub))
                # Completa com cada número do universo que não está no subconjunto
                for x in range(1, self.config.universe_size + 1):
                    # verifica se x já não está no sub
                    if not (base_mask & (1 << (x - 1))):
                        candidates.add(base_mask | (1 << (x - 1)))
        
        total_candidates = len(candidates)
        logger.info(f"Total de candidatos únicos a analisar: {total_candidates:,}")
        
        # Dividir candidatos em chunks para os processos
        candidate_list = list(candidates)
        chunk_size = max(1, total_candidates // (num_processes * 10))
        chunks = [candidate_list[i:i + chunk_size] for i in range(0, total_candidates, chunk_size)]
        
        # Adaptador para o worker_task_candidate
        task_args = []
        for chunk in chunks:
            task_args.append((
                chunk,
                self.config.universe_size,
                self.config.numbers_to_pick,
                self.contest_masks,
                self.config.min_hits_to_save,
                self.config.min_occurences,
                self.config.secondary_hits,
                self.config.secondary_min_occurences
            ))
            
        final_data = []
        pool = multiprocessing.Pool(processes=num_processes)
        try:
            for i, batch_res in enumerate(pool.imap_unordered(worker_task_candidate, task_args)):
                if batch_res:
                    final_data.extend(batch_res)
                if (i + 1) % max(1, (len(chunks) // 10)) == 0:
                    logger.info(f"Analizando candidatos... {(i+1)/len(chunks)*100:.1f}%")
        finally:
            pool.close()
            pool.join()
            
        elapsed = time.time() - start_time
        logger.info(f"Backtest Turbo finalizado em {elapsed:.2f} segundos.")
        logger.info(f"Ranking: {len(final_data)} combinações de elite encontradas.")
        if final_data: self.save_results(final_data)
        return final_data

    def save_results(self, data: List[Tuple]):
        if not os.path.exists('resultados'): os.makedirs('resultados')
        
        # Sanitize for filename and table name
        name_file = self.config.name.lower().replace(' ', '_').replace('-', '_')
        sqlite_path = os.path.join('resultados', 'results.db')
        
        # Pré-processa todos os registros
        rows_data = []
        is_pos = self.config.is_positional
        for comb, max_h, h_min, h_sec in data:
            if is_pos:
                # comb já é uma lista para posicional
                nums = ",".join(map(str, comb))
            else:
                # comb é um mask para set-based
                nums = "-".join(map(str, bitmask_to_numbers(comb, self.config.universe_size)))
            
            score = (max_h * 10) + (h_min * 5) + h_sec
            rows_data.append((nums, max_h, h_min, h_sec, score))

        # ─────────────────────────────────────────────
        # EXPORTAÇÃO EXCEL (openpyxl)
        # ─────────────────────────────────────────────
        if OPENPYXL_AVAILABLE:
            xlsx_path = os.path.join('resultados', f"{name_file}_results.xlsx")
            logger.info(f"Gerando Excel formatado: {xlsx_path}")

            wb = Workbook()
            ws = wb.active
            ws.title = self.config.name[:31]  # max 31 chars

            # ── Paleta de cores ──────────────────────
            COR_HEADER_BG   = "1A237E"  # azul profundo
            COR_HEADER_FONT = "FFFFFF"  # branco
            COR_ROW_ODD     = "E8EAF6"  # azul lavanda claro
            COR_ROW_EVEN    = "FFFFFF"  # branco
            COR_BORDER      = "9FA8DA"  # azul aço
            COR_SCORE_BG    = "283593"  # azul médio (coluna score)
            COR_SCORE_FONT  = "FFD600"  # dourado (coluna score)

            # ── Estilos reutilizáveis ────────────────
            thin = Side(style='thin', color=COR_BORDER)
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            header_fill = PatternFill("solid", fgColor=COR_HEADER_BG)
            header_font = Font(name="Calibri", bold=True, color=COR_HEADER_FONT, size=11)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

            score_fill = PatternFill("solid", fgColor=COR_SCORE_BG)
            score_font = Font(name="Calibri", bold=True, color=COR_SCORE_FONT, size=11)

            fill_odd  = PatternFill("solid", fgColor=COR_ROW_ODD)
            fill_even = PatternFill("solid", fgColor=COR_ROW_EVEN)
            cell_font = Font(name="Calibri", size=10)
            cell_align = Alignment(horizontal="center", vertical="center")

            # ── Linha de título mesclada (linha 1) ───
            title_text = f"🏆  LotteryLab — {self.config.name}  |  Elite Ranking"
            n_dez = self.config.numbers_to_pick
            n_cols = n_dez + 4
            last_col_letter = get_column_letter(n_cols)
            ws.merge_cells(f"A1:{last_col_letter}1")
            title_cell = ws["A1"]
            title_cell.value = title_text
            title_cell.font  = Font(name="Calibri", bold=True, color=COR_HEADER_FONT, size=13)
            title_cell.fill  = PatternFill("solid", fgColor="0D1B8E")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            # ── Linha de subtítulo (linha 2) ─────────
            subtitle_cols = [
                (1, f"Loteria: {self.config.name}"),
                (n_dez + 1, f"Universo: 1-{self.config.universe_size}"),
                (n_dez + 2, f"Números/jogo: {self.config.numbers_to_pick}"),
                (n_dez + 3, f"Mín. acertos: {self.config.min_hits_to_save}"),
                (n_cols, f"Total: {len(rows_data):,} combos"),
            ]
            sub_fill = PatternFill("solid", fgColor="3949AB")
            sub_font = Font(name="Calibri", italic=True, color="E8EAF6", size=9)
            
            # Limpa background e borda de toda a linha de subtítulo
            for c_idx in range(1, n_cols + 1):
                c = ws.cell(row=2, column=c_idx)
                c.fill = sub_fill
                c.border = border
            
            # Mescla as colunas estreitas para o primeiro subtítulo ter espaço suficiente
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_dez)
                
            for col_idx, val in subtitle_cols:
                c = ws.cell(row=2, column=col_idx)
                c.value = val
                c.font  = sub_font
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 18

            # ── Cabeçalho das colunas (linha 3) ──────
            HEADERS = []
            for i in range(1, n_dez + 1):
                HEADERS.append((i, f"D{i}"))
                
            HEADERS.extend([
                (n_dez + 1, "🔝  Acertos Máx."),
                (n_dez + 2, "✅  Hits Principal"),
                (n_dez + 3, "🔸  Hits Secundário"),
                (n_dez + 4, "⭐  Score"),
            ])
            
            for col_idx, label in HEADERS:
                c = ws.cell(row=3, column=col_idx)
                c.value     = label
                c.font      = header_font if col_idx != n_cols else score_font
                c.fill      = header_fill if col_idx != n_cols else score_fill
                c.alignment = header_align
                c.border    = border
            ws.row_dimensions[3].height = 22

            # ── Dados (a partir da linha 4) ───────────
            for row_idx, (nums, max_h, h_min, h_sec, score) in enumerate(rows_data, start=4):
                fill = fill_odd if (row_idx % 2 == 0) else fill_even
                
                # Cada dezena na sua própria célula
                sep = '-' if '-' in nums else ','
                dez_list = [int(n.strip()) for n in nums.split(sep) if n.strip().isdigit()]
                
                # Corrige tamanho da lista preenchendo com 0 caso a extração não seja exata
                if len(dez_list) < n_dez:
                    dez_list.extend([''] * (n_dez - len(dez_list)))
                elif len(dez_list) > n_dez:
                    dez_list = dez_list[:n_dez]
                
                row_vals = dez_list + [max_h, h_min, h_sec, score]
                
                for col_idx, val in enumerate(row_vals, start=1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.fill      = fill
                    c.font      = cell_font
                    c.alignment = cell_align
                    c.border    = border
                ws.row_dimensions[row_idx].height = 16

            # ── Larguras das colunas ──────────────────
            # Largura reduzida para as colunas com números (D1, D2...)
            for i in range(1, n_dez + 1):
                ws.column_dimensions[get_column_letter(i)].width = 5.5
            
            ws.column_dimensions[get_column_letter(n_dez + 1)].width = 18
            ws.column_dimensions[get_column_letter(n_dez + 2)].width = 18
            ws.column_dimensions[get_column_letter(n_dez + 3)].width = 20
            ws.column_dimensions[get_column_letter(n_dez + 4)].width = 12

            # ── Congela painel após cabeçalho ─────────
            ws.freeze_panes = "A4"

            # ── Auto-filtro nos dados ─────────────────
            ws.auto_filter.ref = f"A3:{last_col_letter}{3 + len(rows_data)}"

            wb.save(xlsx_path)
            logger.info(f"Excel formatado salvo: {xlsx_path}")
            
            # ── Criação automática de Backup para Prova de Fogo ──
            try:
                import shutil
                conn_bkp = sqlite3.connect(self.db_path)
                c_bkp = conn_bkp.cursor()
                try:
                    c_bkp.execute(f"SELECT MAX(concurso) FROM {self.config.table_name}")
                    max_c = c_bkp.fetchone()[0] or len(self.contest_masks)
                except Exception:
                    max_c = len(self.contest_masks)
                conn_bkp.close()
                
                target_contest = max_c + 1
                bkp_dir = os.path.join('resultados', 'Bkp_alvo')
                if not os.path.exists(bkp_dir):
                    os.makedirs(bkp_dir)
                    
                bkp_path = os.path.join(bkp_dir, f"{name_file}_results_BKP_Alvo_{target_contest}.xlsx")
                shutil.copy2(xlsx_path, bkp_path)
                logger.info(f"Backup automático criado para o sorteio futuro: {bkp_path}")
                
                # ── Geração do Backup TXT ──
                txt_path = os.path.join(bkp_dir, f"{name_file}_results_BKP_Alvo_{target_contest}.txt")
                with open(txt_path, 'w', encoding='utf-8') as f_txt:
                    meses = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
                    for idx, data_row in enumerate(rows_data):
                        n_str = data_row[0]
                        sep = '-' if '-' in n_str else ','
                        nums_list = [n.strip().zfill(2) for n in n_str.split(sep) if n.strip().isdigit()]
                        
                        if name_file == 'dia_de_sorte':
                            nums_list.append(meses[idx % 12])
                            
                        f_txt.write(" ".join(nums_list) + "\n")
                logger.info(f"Backup TXT criado: {txt_path}")
            except Exception as e:
                logger.warning(f"Erro ao criar backup automático: {e}")

        else:
            # Fallback CSV caso openpyxl não esteja disponível
            import csv
            csv_path = os.path.join('resultados', f"{name_file}_results.csv")
            logger.warning("openpyxl não encontrado. Exportando CSV simples.")
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['numeros', 'acertos_maximos', 'hits_main', 'hits_sec', 'score'])
                writer.writerows(rows_data)

        # ─────────────────────────────────────────────
        # EXPORTAÇÃO SQLite
        # ─────────────────────────────────────────────
        logger.info(f"Atualizando banco SQLite: {sqlite_path}")
        conn = sqlite3.connect(sqlite_path)
        cursor = conn.cursor()
        cursor.execute(f'DROP TABLE IF EXISTS "{name_file}"')
        cursor.execute(f'CREATE TABLE "{name_file}" (numeros TEXT, acertos_maximos INTEGER, hits_main INTEGER, hits_sec INTEGER, score REAL)')
        cursor.executemany(f'INSERT INTO "{name_file}" VALUES (?, ?, ?, ?, ?)', rows_data)
        conn.commit()
        conn.close()
        logger.info(f"Base de dados de elite atualizada com sucesso para: {name_file}")
        logger.info("Processo 100% concluído. Verifique a pasta /resultados.")
